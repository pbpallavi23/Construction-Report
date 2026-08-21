from __future__ import annotations

import io
import os
import re
from typing import Any

from openpyxl import load_workbook






_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "..", "resources", "incident_report_template.xlsx"
)
_SHEET_NAME = "Incident Report"




_CELL_MAP: dict[str, str] = {
    "project": "C2",
    "job_no": "H2",
    "prepared_by": "B3",
    "received_by": "E3",
    "report_date": "H3",
    "team": "B4",
    "distribution": "E4",
    "incident_type": "E6",
    "category": "E7",
    "person_name": "E9",
    "person_address": "E10",
    "employer_name": "E11",
    "employer_address": "E12",
    "employer_phone": "E13",
    "site_address": "E15",
    "incident_location": "E16",
    "incident_date": "E17",
    "incident_time": "E18",
    "description": "A21",
    "injured_body_part": "E25",
    "was_injured": "B27",
    "accident_book_completed": "E27",
    "riddor_reportable": "H27",
    "time_lost": "B29",
    "time_lost_days": "E29",
    "went_to_hospital": "B31",
    "hospital_details": "E31",
    "saw_gp": "B32",
    "gp_details": "E32",
    "action_taken": "E34",
    "further_action": "E35",
    "reporter_name": "D37",
    "report_handed_date": "H37",
    "reporter_position": "D38",
    "signed_off_by_name": "D41",
    "signed_off_date": "H41",
    "signed_off_by_position": "D42",
}



_INCIDENT_TYPE_LABELS = {
    "accident": "Accidents",
    "near_miss": "Near_Miss",
    "dangerous_occurrence": "Dangerous_Occurrence",
}



_YES_NO_FIELDS = {
    "was_injured",
    "accident_book_completed",
    "riddor_reportable",
    "time_lost",
    "went_to_hospital",
    "saw_gp",
}


_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _to_uk_date(value: str) -> str:
    """The form uses DD/MM/YYYY throughout; report_date is set internally as
    an ISO 'YYYY-MM-DD' default, so normalise it here rather than in every
    caller."""
    if _ISO_DATE_RE.match(value):
        year, month, day = value.split("-")
        return f"{day}/{month}/{year}"
    return value


def _display_value(key: str, value: Any) -> Any:
    if value in (None, ""):
        return None
    if key == "incident_type":
        return _INCIDENT_TYPE_LABELS.get(value, value)
    if key in _YES_NO_FIELDS and isinstance(value, str):
        return value.strip().capitalize()
    if key == "report_date" and isinstance(value, str):
        return _to_uk_date(value)
    return value


def export_to_xlsx(report: dict[str, Any]) -> bytes:
    """Fills the official template with one saved incident report's field
    values and returns the resulting workbook as bytes, ready to send as a
    file download. Blank fields are left blank in the exported form too -
    this mirrors exactly what is saved in the database, nothing is
    invented."""
    wb = load_workbook(_TEMPLATE_PATH)
    ws = wb[_SHEET_NAME]

    fields = report.get("fields", {})
    for key, cell in _CELL_MAP.items():
        value = _display_value(key, fields.get(key))
        if value is not None:
            ws[cell] = value

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def export_filename(report: dict[str, Any]) -> str:
    date_part = (report.get("report_date") or "undated").replace("/", "-")
    return f"Incident_Report_{report.get('id', 'report')}_{date_part}.xlsx"
