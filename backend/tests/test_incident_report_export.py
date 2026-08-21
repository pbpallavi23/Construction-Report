import unittest

from tests import support

from openpyxl import load_workbook

from app.services import incident_report_export


def _sample_report(**field_overrides):
    fields = {
        "project": "Riverside Tower",
        "job_no": "J-1001",
        "report_date": "2026-01-15",
        "incident_type": "near_miss",
        "was_injured": "no",
        "description": "Scaffolding plank slipped, no injury.",
    }
    fields.update(field_overrides)
    return {
        "id": "incident-abc123",
        "report_date": "2026-01-15",
        "fields": fields,
    }


class IncidentReportExportTests(unittest.TestCase):
    def test_export_produces_valid_xlsx_bytes(self):
        xlsx_bytes = incident_report_export.export_to_xlsx(_sample_report())
        self.assertTrue(xlsx_bytes.startswith(b"PK"))

    def test_export_writes_field_values_into_expected_cells(self):
        xlsx_bytes = incident_report_export.export_to_xlsx(_sample_report())
        wb = load_workbook(__import__("io").BytesIO(xlsx_bytes))
        ws = wb["Incident Report"]
        self.assertEqual(ws["C2"].value, "Riverside Tower")
        self.assertEqual(ws["H2"].value, "J-1001")

    def test_export_normalises_incident_type_label(self):
        xlsx_bytes = incident_report_export.export_to_xlsx(
            _sample_report(incident_type="dangerous_occurrence")
        )
        wb = load_workbook(__import__("io").BytesIO(xlsx_bytes))
        ws = wb["Incident Report"]
        self.assertEqual(ws["E6"].value, "Dangerous_Occurrence")

    def test_export_converts_iso_date_to_uk_format(self):
        xlsx_bytes = incident_report_export.export_to_xlsx(_sample_report())
        wb = load_workbook(__import__("io").BytesIO(xlsx_bytes))
        ws = wb["Incident Report"]
        self.assertEqual(ws["H3"].value, "15/01/2026")

    def test_export_capitalises_yes_no_fields(self):
        xlsx_bytes = incident_report_export.export_to_xlsx(_sample_report())
        wb = load_workbook(__import__("io").BytesIO(xlsx_bytes))
        ws = wb["Incident Report"]

        self.assertEqual(ws["B27"].value, "No")

    def test_export_leaves_blank_fields_blank(self):
        report = _sample_report()
        report["fields"]["hospital_details"] = None
        xlsx_bytes = incident_report_export.export_to_xlsx(report)
        wb = load_workbook(__import__("io").BytesIO(xlsx_bytes))
        ws = wb["Incident Report"]
        self.assertIsNone(ws["E31"].value)

    def test_export_filename_includes_id_and_date(self):
        filename = incident_report_export.export_filename(_sample_report())
        self.assertIn("incident-abc123", filename)
        self.assertIn("2026-01-15", filename)
        self.assertTrue(filename.endswith(".xlsx"))


if __name__ == "__main__":
    unittest.main()
